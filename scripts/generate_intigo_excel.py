"""
scripts/generate_intigo_excel.py
─────────────────────────────────
Fills Intigo's template_colis.xlsx with order data and writes the result
to a new file. All 3 sheets (Colis, Lieux, Instructions) are preserved.

Usage:
    python3 generate_intigo_excel.py <orders.json> <template.xlsx> <output.xlsx>
"""

import sys
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate(orders_path: str, template_path: str, output_path: str) -> None:
    with open(orders_path, encoding='utf-8') as f:
        orders = json.load(f)

    # Load template — preserves Lieux + Instructions sheets and all header styling
    wb = load_workbook(template_path)
    ws = wb['Colis']

    data_font      = Font(name='Calibri', size=11, color='FF000000')
    white_fill     = PatternFill('solid', fgColor='FFFFFFFF')
    alt_fill       = PatternFill('solid', fgColor='FFF5F5F5')
    data_alignment = Alignment(vertical='center', wrap_text=False)

    for i, order in enumerate(orders):
        row      = i + 2   # row 1 = headers, data starts at row 2
        customer = order.get('customer') or {}
        addr     = order.get('shipping_address') or {}
        items    = order.get('items') or []

        # ── Recipient name ───────────────────────────────────────────────────
        nom = f"{customer.get('first_name', '')} {customer.get('last_name', '')}".strip()

        # ── Phone — strip non-digits, keep last 8 ────────────────────────────
        phone_raw = customer.get('phone') or ''
        phone     = ''.join(c for c in phone_raw if c.isdigit())[-8:]

        # ── Address ──────────────────────────────────────────────────────────
        street   = addr.get('street', '')
        # Intigo calls "governorate" a "ville" — important mapping
        ville    = (addr.get('governorate') or addr.get('city') or addr.get('state') or '').strip()
        district = (addr.get('district') or '').strip()

        # ── Description from items ────────────────────────────────────────────
        parts = []
        for item in items:
            pack    = item.get('pack')
            product = item.get('product')
            qty     = item.get('quantity') or 1
            if pack and isinstance(pack, dict):
                name = pack.get('name') or 'Pack'
            elif product and isinstance(product, dict):
                name = product.get('name') or 'Article'
            else:
                name = 'Article'
            parts.append(f"{name} x{qty}" if qty > 1 else name)
        description = ', '.join(parts) or 'Colis'
        # Intigo max 500 chars
        if len(description) > 500:
            description = description[:497] + '...'

        # ── Order reference ───────────────────────────────────────────────────
        order_id = order.get('_id', '')
        ref      = f"Cmd #{order_id[-6:].upper()}" if order_id else ''

        # ── Amount ────────────────────────────────────────────────────────────
        try:
            montant = round(float(order.get('total_amount') or 0), 3)
        except (TypeError, ValueError):
            montant = 0.0

        row_values = [
            nom,          # A  nom_destinataire   REQUIRED
            phone,        # B  telephone           REQUIRED
            '',           # C  telephone2          optional
            street,       # D  adresse             REQUIRED
            ville,        # E  ville               REQUIRED
            district,     # F  district            REQUIRED
            '',           # G  quartier            optional
            montant,      # H  montant             optional (0 = prépayé)
            1,            # I  taille_colis        optional (1=Petit default)
            0,            # J  ouvrir_colis        optional (0=Non default)
            description,  # K  description_produit optional
            ref,          # L  info_supplementaire optional
        ]

        fill = alt_fill if i % 2 == 1 else white_fill

        for col_idx, value in enumerate(row_values, start=1):
            cell           = ws.cell(row=row, column=col_idx, value=value)
            cell.font      = data_font
            cell.fill      = fill
            cell.alignment = data_alignment

        ws.row_dimensions[row].height = 18

    wb.save(output_path)


if __name__ == '__main__':
    if len(sys.argv) != 4:
        print("Usage: generate_intigo_excel.py <orders.json> <template.xlsx> <output.xlsx>")
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2], sys.argv[3])